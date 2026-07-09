"""사내 자료실(회사 참고자료) — 스프레드시트/문서에서 텍스트를 추출하고,
직원 질문에 맞는 시트/구획을 검색해 돌려준다.

- 표 형식 자료(배관 규격·중량·플랜지 등)는 시트 단위로 쪼개 저장/검색한다.
- 전체를 시스템 프롬프트에 넣으면 토큰이 과도하므로(수만 토큰), 대화 중 AI가
  search_company_docs 도구로 관련 시트만 가져와 근거로 답한다.
"""
import io
import re

MAX_REF_CHARS = 200_000          # 문서당 저장 텍스트 상한
_SHEET_MARK = "## 시트: "


def _fmt_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:g}"           # 12.0 → "12", 3.14 → "3.14"
    return str(v).replace("\n", " ").strip()


def _xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(_SHEET_MARK + str(ws.title).strip())
        for row in ws.iter_rows(values_only=True):
            cells = [_fmt_cell(c) for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if any(cells):
                out.append(" | ".join(cells))
    wb.close()
    return "\n".join(out)


def _xls_text(data: bytes) -> str:
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    out = []
    for sh in wb.sheets():
        out.append(_SHEET_MARK + str(sh.name).strip())
        for i in range(sh.nrows):
            cells = [_fmt_cell(sh.cell_value(i, j)) for j in range(sh.ncols)]
            while cells and cells[-1] == "":
                cells.pop()
            if any(cells):
                out.append(" | ".join(cells))
    return "\n".join(out)


def extract_reference_text(data: bytes, filename: str, content_type: str) -> str | None:
    """참고자료 텍스트 추출. 스프레드시트는 시트별로, 그 외는 기본 추출기로."""
    ext = (filename.rsplit(".", 1)[-1].lower() if "." in filename else "")
    try:
        if ext == "xlsx" or "spreadsheetml" in (content_type or ""):
            return _xlsx_text(data)[:MAX_REF_CHARS]
        if ext == "xls" or content_type == "application/vnd.ms-excel":
            return _xls_text(data)[:MAX_REF_CHARS]
    except Exception:
        return None
    # PDF/텍스트 등은 기존 추출기 재사용
    from app.file_utils import extract_text
    t = extract_text(data, filename, content_type)
    return t[:MAX_REF_CHARS] if t else None


def sheet_titles(text: str) -> list[str]:
    """추출 텍스트에서 시트/구획 제목 목록(자료실 색인 표시용)."""
    return [ln[len(_SHEET_MARK):].strip()
            for ln in (text or "").splitlines() if ln.startswith(_SHEET_MARK)]


def _split_chunks(text: str):
    """(제목, 본문) 청크 목록으로 분리. 시트 마커가 없으면 문단 단위로 자른다."""
    if _SHEET_MARK in (text or ""):
        chunks = []
        title, buf = None, []
        for ln in text.splitlines():
            if ln.startswith(_SHEET_MARK):
                if title is not None:
                    chunks.append((title, "\n".join(buf)))
                title, buf = ln[len(_SHEET_MARK):].strip(), []
            else:
                buf.append(ln)
        if title is not None:
            chunks.append((title, "\n".join(buf)))
        return chunks
    # 마커 없는 문서: ~1200자 단위로 분할
    paras, cur = [], []
    for ln in (text or "").splitlines():
        cur.append(ln)
        if sum(len(x) for x in cur) > 1200:
            paras.append(("", "\n".join(cur))); cur = []
    if cur:
        paras.append(("", "\n".join(cur)))
    return paras


_TOKEN_RE = re.compile(r"[가-힣]{2,}|[A-Za-z]{2,}|[0-9]+(?:/[0-9]+)?(?:\.[0-9]+)?")
# 질문 상용어만 제거(도메인어 무게·인치·규격 등은 검색에 필요하므로 남긴다)
_STOP = {"알려", "얼마", "무엇", "어떻게", "무슨", "인가요", "인지", "대해", "해줘",
         "알려줘", "좀", "그리고", "우리", "회사", "관련", "궁금", "질문"}


# 자료가 한/영 혼용(헤더는 영문, 질문은 한글)이라 도메인어 동의어로 다리를 놓는다.
_SYNONYMS = {
    "볼트": ["bolt"], "스패너": ["spanner"], "사이즈": ["size"], "치수": ["size", "dimension"],
    "무게": ["weight", "wt"], "중량": ["weight", "wt"], "인치": ["inch"],
    "외경": ["od"], "내경": ["id"], "두께": ["thk", "thickness", "wt"],
    "플랜지": ["flange"], "밸브": ["valve"], "게이트": ["gate"], "글로브": ["globe"],
    "체크": ["check"], "강관": ["steel", "pipe"], "동관": ["copper"],
    "스테인리스": ["stainless", "sts"], "스텐": ["stainless", "sts"],
    "인장강도": ["tensile"], "규격": ["spec", "standard"], "구멍": ["hole", "holes"],
    "호칭": ["nps", "nominal"], "관": ["pipe"], "파이프": ["pipe"], "볼팅": ["bolting"],
    "팽창": ["expansion"], "재질": ["material"], "용접": ["welding", "wn"],
}


def _tokens(s: str) -> list[str]:
    toks = [t.lower() for t in _TOKEN_RE.findall(s or "")]
    out = list(toks)
    for t in toks:
        out.extend(_SYNONYMS.get(t, []))
    return out


def search_reference(docs, query: str, max_chars: int = 11000) -> str:
    """docs: [(filename, text)]. 질의어와 겹치는 시트/구획을 점수순으로 골라
    출처 라벨과 함께 텍스트 블록으로 반환. 매칭이 없으면 자료 색인을 돌려준다."""
    qterms = [t for t in _tokens(query) if t not in _STOP]
    scored = []
    for fname, text in docs:
        for title, body in _split_chunks(text):
            hay = (title + "\n" + body).lower()
            if not hay.strip():
                continue
            score = 0
            for t in set(qterms):
                c = hay.count(t)
                if c:
                    score += min(c, 5) + (3 if t in title.lower() else 0)
            if score > 0:
                scored.append((score, fname, title, body))
    scored.sort(key=lambda x: x[0], reverse=True)

    if not scored:
        idx = "\n".join(f"- {f} (구획: {', '.join(sheet_titles(t)[:8]) or '문서'})" for f, t in docs)
        return ("질의와 직접 일치하는 항목을 찾지 못했습니다. 보유 자료 목록은 아래와 같으니 "
                "질문을 구체화하거나 자료명을 참고하세요.\n" + idx)

    out, used = [], 0
    for score, fname, title, body in scored:
        block = f"[출처: {fname}{(' · ' + title) if title else ''}]\n{body.strip()}"
        if used + len(block) > max_chars and out:
            break
        out.append(block[:max_chars])
        used += len(block)
        if used >= max_chars:
            break
    return "\n\n".join(out)
